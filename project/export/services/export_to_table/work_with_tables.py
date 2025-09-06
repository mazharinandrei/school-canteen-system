from string import ascii_uppercase
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side


def center_aligment():
    return Alignment(horizontal="center", vertical="center")


def bold_font():
    return Font(bold=True)


def regular_font():
    return Font(bold=False, italic=False)


def thin_border():
    return Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))


def with_workbook(func):
    """
    Декоратор, создающий openpyxl.Workbook, передающий ws в функцию и сохраняющий файл.
    Функция возвращает путь, по которому сохранился файл
    """

    @wraps(func)
    def wrapper(*args, **kwargs):
        wb = Workbook()
        ws = wb.active
        try:
            filepath = func(ws, *args, **kwargs)
            if not isinstance(filepath, str):
                raise ValueError("Функция должна вернуть строку — путь для сохранения файла.")
            wb.save(filepath)
            return filepath
        finally:
            wb.close()

    return wrapper


def get_columns(first_column_index=0, count=26):
    if count <= 26:
        return list(ascii_uppercase[first_column_index:count + first_column_index])

    # TODO: добавить поддержку большего количества стобцов (AA, AB, AC и тд)


def insert_row(ws, data, row_index=1, widths=None, alignment=None, fonts=None, border=None):
    columns = get_columns(count=len(data))
    for i in range(len(data)):
        cell = f"{columns[i]}{row_index}"
        ws[cell] = data[i]
        if alignment:
            ws[f"{columns[i]}{row_index}"].alignment = alignment

        if widths:
            ws.column_dimensions[ws[cell].column_letter].width = widths[i]

        if fonts:
            ws[f"{columns[i]}{row_index}"].font = fonts[i]

        if border:
            ws[f"{columns[i]}{row_index}"].border = border


